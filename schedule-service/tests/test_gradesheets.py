import io

import openpyxl
import pytest

from app.errors import ScheduleParseError
from app.gradesheets import MAX_ROWS, MAX_SHEETS, parse_gradesheets

HEADER_ROW = 5
HEADER = ("№ п/п", "Фамилия", "Имя", "Отчество")
DEFAULT_BLOCKS = (("М1", ()), ("Зачёт", ("Балл", "Оценка", "Дата", "Преподаватель")))


def dump(workbook) -> bytes:
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_book(
    rows,
    *,
    title="ИДБ-25-11",
    group="ИДБ-25-11",
    discipline="Физика",
    teachers="Иванов И.И., Петров П.П.",
    blocks=DEFAULT_BLOCKS,
    header=HEADER,
) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = title
    if group:
        sheet["B1"] = "Группа:"
        sheet["C1"] = group
    if discipline:
        sheet["B2"] = "Дисциплина:"
        sheet["C2"] = discipline
    sheet["B3"] = "Преподаватель:"
    sheet["C3"] = teachers
    for column, name in enumerate(header, start=1):
        sheet.cell(row=HEADER_ROW, column=column, value=name)
    column = len(header) + 1
    for name, subcolumns in blocks:
        sheet.cell(row=HEADER_ROW, column=column, value=name)
        if subcolumns:
            sheet.merge_cells(
                start_row=HEADER_ROW,
                start_column=column,
                end_row=HEADER_ROW,
                end_column=column + len(subcolumns) - 1,
            )
            for offset, subtitle in enumerate(subcolumns):
                sheet.cell(row=HEADER_ROW + 1, column=column + offset, value=subtitle)
        column += max(len(subcolumns), 1)
    for index, values in enumerate(rows, start=HEADER_ROW + 2):
        for column_index, value in enumerate(values, start=1):
            if value is not None:
                sheet.cell(row=index, column=column_index, value=value)
    return dump(workbook)


def student_of(book: bytes, position: int = 0):
    return parse_gradesheets(book).sheets[0].students[position]


@pytest.fixture(scope="module")
def parsed(gradesheet_xlsx):
    return parse_gradesheets(gradesheet_xlsx)


@pytest.fixture(scope="module")
def first_sheet(parsed):
    return parsed.sheets[0]


@pytest.fixture(scope="module")
def by_last_name(first_sheet):
    return {student.last_name: student for student in first_sheet.students}


class TestRealGradeSheet:
    def test_every_sheet_of_the_book_is_parsed(self, parsed):
        assert [sheet.sheet_name for sheet in parsed.sheets] == [
            "ИДБ-25-11",
            "ИДБ-25-12",
            "ИДБ-25-13",
            "ИДБ-25-14",
            "ИДБ-25-15",
        ]

    def test_header_of_the_sheet(self, first_sheet):
        assert first_sheet.group == "ИДБ-25-11"
        assert first_sheet.discipline == "Технические средства информационных систем"
        assert first_sheet.department == "УИТС"
        assert first_sheet.teachers == ["Чеканин В.А."]
        assert first_sheet.semester == "Весенний семестр 2025/2026 учебного года"
        assert first_sheet.direction == "09.03.03 «Прикладная информатика»"

    def test_blocks_come_from_the_header_in_order(self, first_sheet):
        assert first_sheet.blocks == [
            "М1",
            "М2",
            "Курсовой проект",
            "Зачёт",
            "Повторный зачёт (1)",
            "Повторный зачёт (2)",
        ]

    def test_student_count_per_sheet(self, parsed):
        assert [len(sheet.students) for sheet in parsed.sheets] == [29, 30, 29, 29, 31]

    def test_first_student_matches_contract(self, first_sheet):
        assert first_sheet.students[0].model_dump() == {
            "number": 1,
            "last_name": "Абрамов",
            "first_name": "Александр",
            "patronymic": "Абдул-Керимович",
            "marks": [
                {
                    "block": "М1",
                    "score": 30.0,
                    "text": None,
                    "grade": None,
                    "date": None,
                    "teacher": None,
                },
                {
                    "block": "М2",
                    "score": 29.0,
                    "text": None,
                    "grade": None,
                    "date": None,
                    "teacher": None,
                },
                {
                    "block": "Зачёт",
                    "score": 30.0,
                    "text": None,
                    "grade": "зачтено",
                    "date": "2026-06-02",
                    "teacher": "Чеканин В.А.",
                },
            ],
        }

    def test_template_tail_rows_are_not_students(self, first_sheet):
        assert all(student.last_name for student in first_sheet.students)
        assert first_sheet.students[-1].number == 29

    def test_student_after_the_alphabetical_end_is_kept(self, by_last_name):
        assert by_last_name["Аверин"].number == 29
        assert by_last_name["Аверин"].marks == []

    def test_dashes_mean_no_data(self, by_last_name):
        assert by_last_name["Волков"].marks == []

    def test_text_instead_of_a_score(self, by_last_name):
        credit = by_last_name["Галкин"].marks[-1]
        assert credit.block == "Зачёт"
        assert credit.score is None
        assert credit.text == "не допущен"
        assert credit.date == "2026-06-06"

    def test_missing_patronymic_is_a_warning_not_an_error(self, first_sheet, by_last_name):
        assert by_last_name["Виноградов"].patronymic is None
        assert first_sheet.warnings == [
            "у студента 'Виноградов' в строке 29 пустое отчество"
        ]

    def test_group_in_the_header_may_disagree_with_the_sheet_name(self, parsed):
        assert parsed.sheets[4].warnings == [
            "группа в шапке 'ИДБ-25-14' не совпадает с именем листа 'ИДБ-25-15'"
        ]

    def test_two_teachers_in_one_cell(self, parsed):
        assert parsed.sheets[4].teachers == ["Чеканин В.А.", "Ступивцев А.В."]

    def test_formula_without_a_cached_value_does_not_break_the_header(self, first_sheet):
        assert first_sheet.semester is not None


class TestColumnLayout:
    def test_block_without_subcolumns_is_read_as_a_score(self):
        student = student_of(build_book([[1, "Петров", "Пётр", "Петрович", 42]]))
        assert student.marks[0].block == "М1"
        assert student.marks[0].score == 42.0

    def test_unknown_subcolumn_is_reported_and_skipped(self):
        book = build_book(
            [[1, "Петров", "Пётр", "Петрович", 30, "чем-то ещё"]],
            blocks=(("Зачёт", ("Балл", "Комментарий")),),
        )
        sheet = parse_gradesheets(book).sheets[0]
        assert sheet.warnings == [
            "в блоке 'Зачёт' неизвестная колонка 'комментарий', её значения пропущены"
        ]
        mark = sheet.students[0].marks[0]
        assert mark.score == 30.0
        assert mark.text is None
        assert mark.grade is None

    def test_column_order_is_taken_from_the_header(self):
        book = build_book(
            [[1, "Петров", "Пётр", "Петрович", "Иванов И.И.", "2026-06-02", "зачтено", 30]],
            blocks=(("Зачёт", ("Преподаватель", "Дата", "Оценка", "Балл")),),
        )
        mark = student_of(book).marks[0]
        assert mark.score == 30.0
        assert mark.grade == "зачтено"
        assert mark.date == "2026-06-02"
        assert mark.teacher == "Иванов И.И."

    def test_row_with_a_non_numeric_position(self):
        student = student_of(build_book([["—", "Петров", "Пётр", "Петрович", 42]]))
        assert student.number is None
        assert student.last_name == "Петров"


class TestValues:
    def test_date_written_as_text(self):
        book = build_book([[1, "Петров", "Пётр", "Петрович", None, 30, "зачтено", "02.06.2026"]])
        assert student_of(book).marks[-1].date == "2026-06-02"

    def test_unparsable_date_is_reported(self):
        book = build_book([[1, "Петров", "Пётр", "Петрович", None, 30, "зачтено", "скоро"]])
        sheet = parse_gradesheets(book).sheets[0]
        assert sheet.students[0].marks[-1].date is None
        assert sheet.warnings == ["в блоке 'Зачёт', строка 7: 'скоро' не разобрано как дата"]

    def test_empty_block_produces_no_mark(self):
        book = build_book([[1, "Петров", "Пётр", "Петрович", None, "-----", "-----"]])
        assert student_of(book).marks == []


class TestRejectedBooks:
    def test_not_an_excel_book(self):
        with pytest.raises(ScheduleParseError, match="книгу Excel"):
            parse_gradesheets(b"definitely not a workbook")

    def test_sheet_without_a_header_row(self):
        with pytest.raises(ScheduleParseError, match="Фамилия"):
            parse_gradesheets(build_book([[1, "Петров"]], header=("№ п/п",)))

    def test_sheet_without_a_group(self):
        with pytest.raises(ScheduleParseError, match="Группа:"):
            parse_gradesheets(build_book([[1, "Петров", "Пётр", "Петрович", 42]], group=""))

    def test_sheet_without_a_discipline(self):
        with pytest.raises(ScheduleParseError, match="Дисциплина:"):
            parse_gradesheets(build_book([[1, "Петров", "Пётр", "Петрович", 42]], discipline=""))

    def test_sheet_without_blocks(self):
        with pytest.raises(ScheduleParseError, match="блока оценок"):
            parse_gradesheets(build_book([[1, "Петров", "Пётр", "Петрович"]], blocks=()))

    def test_sheet_without_students(self):
        with pytest.raises(ScheduleParseError, match="нет ни одной строки с фамилией"):
            parse_gradesheets(build_book([]))

    def test_empty_book(self):
        with pytest.raises(ScheduleParseError, match="нет ни одного листа с ведомостью"):
            parse_gradesheets(dump(openpyxl.Workbook()))

    def test_book_over_the_sheet_limit(self):
        workbook = openpyxl.Workbook()
        for index in range(MAX_SHEETS):
            workbook.create_sheet(f"Лист {index}")
        with pytest.raises(ScheduleParseError, match=f"при пределе {MAX_SHEETS}"):
            parse_gradesheets(dump(workbook))

    def test_sheet_over_the_row_limit(self):
        workbook = openpyxl.Workbook()
        workbook.active.cell(row=MAX_ROWS + 1, column=1, value="хвост")
        with pytest.raises(ScheduleParseError, match=f"при пределе {MAX_ROWS}"):
            parse_gradesheets(dump(workbook))


class TestEmptySheetsAreSkipped:
    def test_empty_sheet_next_to_a_real_one(self, gradesheet_xlsx):
        workbook = openpyxl.load_workbook(io.BytesIO(gradesheet_xlsx))
        workbook.create_sheet("Пустой", 0)
        parsed = parse_gradesheets(dump(workbook))
        assert [sheet.sheet_name for sheet in parsed.sheets] == [
            "ИДБ-25-11",
            "ИДБ-25-12",
            "ИДБ-25-13",
            "ИДБ-25-14",
            "ИДБ-25-15",
        ]
