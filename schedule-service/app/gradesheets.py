import io
import re
import zipfile
from datetime import date, datetime
from typing import Any, BinaryIO, NamedTuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

from app.errors import ScheduleParseError
from app.models import GradeSheet, GradeSheetStudent, Mark, ParsedGradeSheets
from app.parser import GROUP_RE

MAX_SHEETS = 60
MAX_ROWS = 5000
MAX_UNPACKED_BYTES = 64 * 1024 * 1024
HEADER_SEARCH_ROWS = 30

NO_DATA_RE = re.compile(r"^-+$")
SEMESTER_RE = re.compile(r"семестр", re.IGNORECASE)
DIRECTION_RE = re.compile(r"^\d{2}\.\d{2}\.\d{2}\b")
RU_DATE_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$")

NAME_TITLES = {"фамилия": "last_name", "имя": "first_name", "отчество": "patronymic"}
NUMBER_TITLES = {"№ п/п", "№", "№ п.п."}
SCORE, GRADE, DATE, TEACHER = "балл", "оценка", "дата", "преподаватель"
SUBCOLUMN_TITLES = {SCORE, GRADE, DATE, TEACHER}
META_TITLES = {
    "дисциплина": "discipline",
    "группа": "group",
    "кафедра": "department",
    "преподаватель": "teachers",
}


class _Block(NamedTuple):
    title: str
    columns: dict[str, int]
    single: int | None


def parse_gradesheets(source: bytes | BinaryIO) -> ParsedGradeSheets:
    if isinstance(source, bytes):
        source = io.BytesIO(source)
    _check_archive(source)
    try:
        workbook = openpyxl.load_workbook(source, data_only=True)
    except Exception as e:
        raise ScheduleParseError("файл не удалось прочитать как книгу Excel") from e
    try:
        if len(workbook.worksheets) > MAX_SHEETS:
            raise ScheduleParseError(
                f"в книге {len(workbook.worksheets)} листов при пределе {MAX_SHEETS}"
            )
        sheets = [parsed for sheet in workbook.worksheets if (parsed := _parse_sheet(sheet))]
    except ValidationError as e:
        raise ScheduleParseError(f"результат разбора не прошёл валидацию схемы: {e}") from e
    finally:
        workbook.close()
    if not sheets:
        raise ScheduleParseError("в книге нет ни одного листа с ведомостью")
    return ParsedGradeSheets(sheets=sheets)


def _check_archive(source: BinaryIO) -> None:
    try:
        with zipfile.ZipFile(source) as archive:
            unpacked = sum(item.file_size for item in archive.infolist())
    except zipfile.BadZipFile as e:
        raise ScheduleParseError("файл не удалось прочитать как книгу Excel") from e
    if unpacked > MAX_UNPACKED_BYTES:
        limit_mb = MAX_UNPACKED_BYTES // (1024 * 1024)
        raise ScheduleParseError(
            f"распакованное содержимое книги больше допустимых {limit_mb} МБ"
        )
    source.seek(0)


def _parse_sheet(sheet: Worksheet) -> GradeSheet | None:
    if sheet.max_row > MAX_ROWS:
        raise ScheduleParseError(
            f"в листе '{sheet.title}' {sheet.max_row} строк при пределе {MAX_ROWS}"
        )
    header = _find_header(sheet)
    if header is None:
        if _is_empty(sheet):
            return None
        raise ScheduleParseError(
            f"в листе '{sheet.title}' не нашлась строка заголовка с колонкой 'Фамилия'"
        )
    header_row, columns = header
    warnings: list[str] = []
    meta = _parse_meta(sheet, header_row)
    for label, field in (("Группа:", "group"), ("Дисциплина:", "discipline")):
        if not meta.get(field):
            raise ScheduleParseError(f"в шапке листа '{sheet.title}' не нашлась строка '{label}'")
    blocks, header_end = _parse_blocks(sheet, header_row, max(columns.values()) + 1, warnings)
    if not blocks:
        raise ScheduleParseError(
            f"в листе '{sheet.title}' правее колонок ФИО нет ни одного блока оценок"
        )
    students = _parse_students(sheet, columns, blocks, header_end + 1, warnings)
    if not students:
        raise ScheduleParseError(f"в листе '{sheet.title}' нет ни одной строки с фамилией")
    group = meta["group"]
    if GROUP_RE.match(sheet.title) and sheet.title != group:
        warnings.append(f"группа в шапке '{group}' не совпадает с именем листа '{sheet.title}'")
    return GradeSheet(
        sheet_name=sheet.title,
        group=group,
        discipline=meta["discipline"],
        department=meta.get("department"),
        teachers=_split_teachers(meta.get("teachers")),
        semester=meta.get("semester"),
        direction=meta.get("direction"),
        blocks=[block.title for block in blocks],
        students=students,
        warnings=warnings,
    )


def _find_header(sheet: Worksheet) -> tuple[int, dict[str, int]] | None:
    for row in range(1, min(sheet.max_row, HEADER_SEARCH_ROWS) + 1):
        columns: dict[str, int] = {}
        for cell in sheet[row]:
            title = _title(cell.value)
            if title in NAME_TITLES:
                columns.setdefault(NAME_TITLES[title], cell.column)
            elif title in NUMBER_TITLES:
                columns.setdefault("number", cell.column)
        if "last_name" in columns:
            return row, columns
    return None


def _is_empty(sheet: Worksheet) -> bool:
    for row in sheet.iter_rows(values_only=True):
        if any(value is not None for value in row):
            return False
    return True


def _parse_meta(sheet: Worksheet, header_row: int) -> dict[str, str]:
    meta: dict[str, str] = {}
    for row in range(1, header_row):
        for cell in sheet[row]:
            title = _title(cell.value)
            if title in META_TITLES:
                value = _value_right_of(sheet, row, cell.column)
                if value:
                    meta.setdefault(META_TITLES[title], value)
                continue
            text = _text(cell.value)
            if not text:
                continue
            if "semester" not in meta and SEMESTER_RE.search(text):
                meta["semester"] = text
            elif "direction" not in meta and DIRECTION_RE.match(text):
                meta["direction"] = text
    return meta


def _value_right_of(sheet: Worksheet, row: int, column: int) -> str | None:
    for next_column in range(column + 1, sheet.max_column + 1):
        value = _text(sheet.cell(row=row, column=next_column).value)
        if value:
            return value
    return None


def _parse_blocks(
    sheet: Worksheet, header_row: int, first_column: int, warnings: list[str]
) -> tuple[list[_Block], int]:
    spans = {
        merged.min_col: merged.max_col
        for merged in sheet.merged_cells.ranges
        if merged.min_row == header_row
    }
    header_end = max(
        [header_row]
        + [
            merged.max_row
            for merged in sheet.merged_cells.ranges
            if merged.min_row == header_row
        ]
    )
    blocks: list[_Block] = []
    for cell in sheet[header_row]:
        if cell.column < first_column:
            continue
        title = _text(cell.value)
        if not title:
            continue
        subcolumns: dict[str, int] = {}
        unknown: list[str] = []
        for column in range(cell.column, spans.get(cell.column, cell.column) + 1):
            subtitle = _title(sheet.cell(row=header_row + 1, column=column).value)
            if subtitle is None:
                continue
            if subtitle in SUBCOLUMN_TITLES:
                subcolumns.setdefault(subtitle, column)
            else:
                unknown.append(subtitle)
        for subtitle in unknown:
            warnings.append(
                f"в блоке '{title}' неизвестная колонка '{subtitle}', её значения пропущены"
            )
        if subcolumns or unknown:
            header_end = max(header_end, header_row + 1)
        blocks.append(
            _Block(
                title=title,
                columns=subcolumns,
                single=cell.column if not subcolumns and not unknown else None,
            )
        )
    return blocks, header_end


def _parse_students(
    sheet: Worksheet,
    columns: dict[str, int],
    blocks: list[_Block],
    first_row: int,
    warnings: list[str],
) -> list[GradeSheetStudent]:
    students: list[GradeSheetStudent] = []
    for row in range(first_row, sheet.max_row + 1):
        last_name = _text(sheet.cell(row=row, column=columns["last_name"]).value)
        if not last_name:
            continue
        patronymic = _column_text(sheet, row, columns.get("patronymic"))
        if not patronymic:
            warnings.append(f"у студента '{last_name}' в строке {row} пустое отчество")
        students.append(
            GradeSheetStudent(
                number=_number(sheet, row, columns.get("number")),
                last_name=last_name,
                first_name=_column_text(sheet, row, columns.get("first_name")),
                patronymic=patronymic,
                marks=[
                    mark for block in blocks if (mark := _parse_mark(sheet, row, block, warnings))
                ],
            )
        )
    return students


def _parse_mark(sheet: Worksheet, row: int, block: _Block, warnings: list[str]) -> Mark | None:
    if block.single is not None:
        score, text = _score_or_text(sheet.cell(row=row, column=block.single).value)
        if score is None and text is None:
            return None
        return Mark(block=block.title, score=score, text=text)
    score, text = _score_or_text(_column_value(sheet, row, block.columns.get(SCORE)))
    grade = _column_text(sheet, row, block.columns.get(GRADE))
    teacher = _column_text(sheet, row, block.columns.get(TEACHER))
    raw_date = _column_value(sheet, row, block.columns.get(DATE))
    mark_date = _date(raw_date)
    if mark_date is None and (unparsed := _text(raw_date)):
        warnings.append(
            f"в блоке '{block.title}', строка {row}: '{unparsed}' не разобрано как дата"
        )
    if score is None and not any((text, grade, teacher, mark_date)):
        return None
    return Mark(
        block=block.title, score=score, text=text, grade=grade, date=mark_date, teacher=teacher
    )


def _column_value(sheet: Worksheet, row: int, column: int | None) -> Any:
    return None if column is None else sheet.cell(row=row, column=column).value


def _column_text(sheet: Worksheet, row: int, column: int | None) -> str | None:
    return _text(_column_value(sheet, row, column))


def _number(sheet: Worksheet, row: int, column: int | None) -> int | None:
    value = _column_value(sheet, row, column)
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    return int(value) if value >= 1 and float(value).is_integer() else None


def _score_or_text(value: Any) -> tuple[float | None, str | None]:
    if isinstance(value, bool):
        return None, None
    if isinstance(value, (int, float)):
        return float(value), None
    return None, _text(value)


def _date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _text(value)
    if not text:
        return None
    match = RU_DATE_RE.match(text)
    try:
        if match:
            day, month, year = (int(group) for group in match.groups())
            return date(year, month, day).isoformat()
        return date.fromisoformat(text).isoformat()
    except ValueError:
        return None


def _split_teachers(value: str | None) -> list[str]:
    if not value:
        return []
    return [teacher for token in value.split(",") if (teacher := token.strip())]


def _title(value: Any) -> str | None:
    text = _text(value)
    return text.lower().rstrip(":").strip() if text else None


def _text(value: Any) -> str | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if not isinstance(value, str):
        return str(value)
    cleaned = re.sub(r"\s+", " ", value).strip()
    return None if not cleaned or NO_DATA_RE.match(cleaned) else cleaned
